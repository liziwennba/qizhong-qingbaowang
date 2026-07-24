#!/usr/bin/env python3
from collections import defaultdict
from datetime import datetime
from openpyxl import load_workbook
import hashlib, json, os, re, sys

if len(sys.argv) < 3:
    print("用法: python convert_xlsx_to_player_data.py 输入.xlsx 输出/player-data.js [输出/player-data.json]")
    sys.exit(1)

xlsx = sys.argv[1]
out_js = sys.argv[2]
out_json = sys.argv[3] if len(sys.argv) >= 4 else os.path.splitext(out_js)[0] + '.json'
wb = load_workbook(xlsx, read_only=True, data_only=True)


def workbook_sheet_contexts(workbook):
    contexts = []
    required_headers = ['大营武将', '中军武将', '前锋武将']
    for ws in workbook.worksheets:
        rows = ws.iter_rows(values_only=True)
        try:
            headers = next(rows)
        except StopIteration:
            continue

        header_map = {str(v).strip(): i for i, v in enumerate(headers) if v is not None}
        missing = [h for h in required_headers if h not in header_map]
        if missing:
            print(f'跳过工作表 {ws.title}: 缺少 {", ".join(missing)}')
            continue

        contexts.append({
            'sheetName': ws.title,
            'rows': rows,
            'nameIdx': header_map.get('名字', 0),
            'redSumIdx': header_map.get('阵容红度'),
            'mainIdx': header_map['大营武将'],
            'middleIdx': header_map['中军武将'],
            'frontIdx': header_map['前锋武将'],
            'mainSkillIdx': header_map.get('大营技能'),
            'middleSkillIdx': header_map.get('中军技能'),
            'frontSkillIdx': header_map.get('前锋技能'),
            'recordTypeIdx': header_map.get('记录类型'),
            'timeIdx': header_map.get('记录时间'),
        })
    return contexts


sheet_contexts = workbook_sheet_contexts(wb)
if not sheet_contexts:
    print('未找到包含队伍字段的工作表')
    sys.exit(1)


def clean_lines(cell):
    if cell is None:
        return []
    s = str(cell).replace('\r', '\n')
    return [p.strip() for p in s.split('\n') if p and str(p).strip()]


def extract_slot(general_cell, skill_cell):
    parts = clean_lines(general_cell)
    skills = clean_lines(skill_cell)
    red = ''
    level = ''
    name = ''
    if parts:
        if len(parts) >= 1 and '红' in parts[0]:
            red = parts[0]
        if len(parts) >= 2 and '级' in parts[1]:
            level = parts[1]
        name = parts[-1]
    return {
        'name': name,
        'red': red,
        'level': level,
        'skills': skills,
    }


def norm_time(value):
    if value is None:
        return ''
    if isinstance(value, datetime):
        return value.strftime('%Y/%m/%d %H:%M:%S')
    s = str(value).strip()
    for fmt in ['%Y/%m/%d %H:%M:%S', '%Y-%m-%d %H:%M:%S', '%Y/%m/%d %H:%M', '%Y-%m-%d %H:%M']:
        try:
            return datetime.strptime(s, fmt).strftime('%Y/%m/%d %H:%M:%S')
        except Exception:
            pass
    return s


def sort_key_time(s):
    try:
        return datetime.strptime(s, '%Y/%m/%d %H:%M:%S')
    except Exception:
        return datetime.min


def source_signature(path):
    h = hashlib.sha256()
    with open(path, 'rb') as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b''):
            h.update(chunk)
    return h.hexdigest()[:16]


def refresh_index_cache_buster(out_js_path, signature):
    docs_dir = os.path.dirname(os.path.abspath(out_js_path))
    index_path = os.path.join(docs_dir, 'index.html')
    if not os.path.exists(index_path):
        return

    script_name = os.path.basename(out_js_path)
    with open(index_path, 'r', encoding='utf-8') as f:
        text = f.read()

    pattern = rf'src="{re.escape(script_name)}(?:\?v=[^"]*)?"'
    updated = re.sub(pattern, f'src="{script_name}?v={signature}"', text)
    if updated != text:
        with open(index_path, 'w', encoding='utf-8') as f:
            f.write(updated)
        print(f'已更新缓存版本: {index_path}')


def safe_value(row, idx):
    if idx is None or idx >= len(row):
        return None
    return row[idx]


def lineup_name_text(main_slot, middle_slot, front_slot):
    return ' / '.join([
        main_slot.get('name') or '-',
        middle_slot.get('name') or '-',
        front_slot.get('name') or '-',
    ])


player_records = defaultdict(list)
record_count = 0
latest_record_time = ''
for ctx in sheet_contexts:
    for row in ctx['rows']:
        if not row:
            continue
        name = row[ctx['nameIdx']] if ctx['nameIdx'] < len(row) else None
        if name is None:
            continue
        name = str(name).strip()
        if not name:
            continue

        main_slot = extract_slot(safe_value(row, ctx['mainIdx']), safe_value(row, ctx['mainSkillIdx']))
        middle_slot = extract_slot(safe_value(row, ctx['middleIdx']), safe_value(row, ctx['middleSkillIdx']))
        front_slot = extract_slot(safe_value(row, ctx['frontIdx']), safe_value(row, ctx['frontSkillIdx']))
        team_red = safe_value(row, ctx['redSumIdx'])
        team_red_text = '' if team_red is None else str(team_red).strip()
        record_type = str(safe_value(row, ctx['recordTypeIdx']) or '队伍表记录').strip() or '队伍表记录'
        t = norm_time(safe_value(row, ctx['timeIdx']))
        record_count += 1
        if sort_key_time(t) > sort_key_time(latest_record_time):
            latest_record_time = t
        player_records[name].append({
            'time': t,
            'recordType': record_type,
            'teamRed': team_red_text,
            'main': main_slot.get('name', ''),
            'middle': middle_slot.get('name', ''),
            'front': front_slot.get('name', ''),
            'mainSlot': main_slot,
            'middleSlot': middle_slot,
            'frontSlot': front_slot,
            'lineupText': lineup_name_text(main_slot, middle_slot, front_slot),
        })

players = {}
player_list = []
for name in sorted(player_records.keys(), key=lambda x: x.lower()):
    recs = sorted(player_records[name], key=lambda r: sort_key_time(r['time']), reverse=True)
    grouped = {}
    order = []
    for r in recs:
        key = (
            r['mainSlot'].get('name', ''), r['mainSlot'].get('red', ''), r['mainSlot'].get('level', ''), tuple(r['mainSlot'].get('skills', [])),
            r['middleSlot'].get('name', ''), r['middleSlot'].get('red', ''), r['middleSlot'].get('level', ''), tuple(r['middleSlot'].get('skills', [])),
            r['frontSlot'].get('name', ''), r['frontSlot'].get('red', ''), r['frontSlot'].get('level', ''), tuple(r['frontSlot'].get('skills', [])),
        )
        if key not in grouped:
            grouped[key] = {
                'lineupText': r['lineupText'],
                'main': r['main'],
                'middle': r['middle'],
                'front': r['front'],
                'mainSlot': r['mainSlot'],
                'middleSlot': r['middleSlot'],
                'frontSlot': r['frontSlot'],
                'teamRed': r['teamRed'],
                'count': 0,
                'latestTime': r['time'],
                'recordTypes': [],
            }
            order.append(key)
        g = grouped[key]
        g['count'] += 1
        if r['recordType'] not in g['recordTypes']:
            g['recordTypes'].append(r['recordType'])
        if sort_key_time(r['time']) > sort_key_time(g['latestTime']):
            g['latestTime'] = r['time']
            g['teamRed'] = r['teamRed']
    lineups = sorted([grouped[k] for k in order], key=lambda x: sort_key_time(x['latestTime']), reverse=True)
    recent_records = [{
        'time': r['time'],
        'recordType': r['recordType'],
        'teamRed': r['teamRed'],
        'lineupText': r['lineupText'],
        'mainSlot': r['mainSlot'],
        'middleSlot': r['middleSlot'],
        'frontSlot': r['frontSlot'],
    } for r in recs[:20]]
    latest = lineups[0] if lineups else None
    players[name] = {
        'name': name,
        'totalRecords': len(recs),
        'lineups': lineups,
        'lineupCount': len(lineups),
        'recentRecords': recent_records,
        'latestLineup': latest,
    }
    player_list.append({
        'name': name,
        'totalRecords': len(recs),
        'lineupCount': len(lineups),
        'latestLineupText': latest['lineupText'] if latest else '',
    })

signature = source_signature(xlsx)
payload = {
    'updatedAt': f'由 {os.path.basename(xlsx)} 自动生成（{record_count} 条记录）',
    'sourceFile': os.path.basename(xlsx),
    'sourceSignature': signature,
    'sourceSheets': [ctx['sheetName'] for ctx in sheet_contexts],
    'sourceRecordCount': record_count,
    'sourceLatestRecordTime': latest_record_time,
    'playerCount': len(players),
    'players': players,
    'playerList': player_list,
}
with open(out_js, 'w', encoding='utf-8') as f:
    f.write('window.PLAYER_DB = ')
    json.dump(payload, f, ensure_ascii=False, separators=(',', ':'))
    f.write(';\n')
print(f'已生成: {out_js}')
with open(out_json, 'w', encoding='utf-8') as f:
    json.dump(payload, f, ensure_ascii=False, indent=2)
    f.write('\n')
print(f'已生成: {out_json}')
refresh_index_cache_buster(out_js, signature)
